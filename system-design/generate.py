"""Generate the Baby Sign Helper system-design PDF and Excel workbook.

Run: py -3.13 system-design/generate.py
Output: system-design/baby-sign-helper-system-design.{pdf,xlsx}
"""
import os
from datetime import date

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak,
    Table, TableStyle, KeepTogether,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# Paths
# ----------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DOCS = os.path.join(REPO, "docs")

PDF_OUT = os.path.join(HERE, "baby-sign-helper-system-design.pdf")
XLSX_OUT = os.path.join(HERE, "baby-sign-helper-system-design.xlsx")

# ----------------------------------------------------------------------
# Data — single source of truth for both formats
# ----------------------------------------------------------------------
META = {
    "title": "Baby Sign Helper",
    "subtitle": "System Design Document",
    "version": "1.0",
    "date": date.today().isoformat(),
    "owner": "RahmanBhuiyan",
    "repo": "github.com/RahmanBhuiyan/baby-sign-language-webapp",
    "based_on": "RoxyDiya/BABY-SIGN-LANGUAGE-RECOGNITION (Sapienza University of Rome)",
}

OVERVIEW = (
    "Baby Sign Helper is a friendly cross-platform application that reads "
    "baby sign-language gestures in real time and translates them into "
    "spoken words. A Vue 3 front-end captures webcam frames every 200 ms "
    "and POSTs them to a Flask backend hosted on Hugging Face Spaces; the "
    "backend uses Google's MediaPipe HandLandmarker to find the hand, then "
    "feeds the cropped region into a fine-tuned ResNet-18 classifier "
    "(12 classes) and returns the top-1 label plus confidence. The same "
    "Vue app is packaged with Capacitor into an Android APK so parents "
    "can install it on a phone and use the app on cellular data."
)

COMPONENTS = [
    # (component, role, technology, runtime location)
    ("Web frontend", "Camera capture + UI rendering, sends JPEG frames to /api/predict, plays back the spoken word",
     "Vue 3 (Composition API) + Vite", "Browser / Android WebView"),
    ("Flask API", "Decodes JPEG, calls hand detector, crops 224x224, runs ResNet, returns top-1 label",
     "Flask 3 + flask-cors + gunicorn", "Docker container on HF Spaces"),
    ("Hand detector", "Locates the hand bbox using 21 landmarks",
     "MediaPipe Tasks HandLandmarker", "Same container as Flask"),
    ("Classifier model", "ResNet-18 backbone with custom 256-unit dropout head over 12 classes",
     "PyTorch 2.x + torchvision", "Same container as Flask"),
    ("Android shell", "Wraps the built Vue app as a native APK; handles getUserMedia permission",
     "Capacitor 7 + Android WebView", "User's Android phone"),
    ("CI/CD", "Builds the Vue app, generates the Android project, signs the debug APK",
     "GitHub Actions (Node 20, JDK 21, Android SDK)", "GitHub-hosted Ubuntu runner"),
    ("Source control", "Two remotes from one working copy",
     "Git", "GitHub (code + workflow) and HF Spaces (Docker deploy)"),
]

TECH_STACK = [
    # (layer, technology, version, purpose)
    ("Frontend framework", "Vue", "3.5.x", "Reactive component model + Composition API"),
    ("Frontend bundler",   "Vite", "5.4.x", "Dev server + production build"),
    ("Mobile shell",       "Capacitor", "7.x", "Wraps the Vue app as Android APK"),
    ("Browser APIs",       "getUserMedia, Web Speech API", "native", "Camera capture + voice playback"),
    ("Backend framework",  "Flask", "3.x",   "HTTP routing"),
    ("WSGI server",        "gunicorn", "22.x", "Production process manager"),
    ("CORS",               "flask-cors", "6.x", "Allow APK / browser cross-origin calls"),
    ("ML framework",       "PyTorch", "2.x", "Inference"),
    ("Computer vision",    "OpenCV", "4.13.x", "Image decoding + colour conversion"),
    ("Hand detection",     "MediaPipe Tasks", "0.10.x", "21-landmark hand keypoint model"),
    ("Image preprocessing","Pillow", "12.x", "JPEG decode"),
    ("Numerics",           "NumPy", "2.x", "Tensor manipulation"),
    ("HTTP client (auto-download)", "requests", "2.x", "Fetch model weights on first boot"),
    ("Container base",     "python:3.11-slim (Debian 12)", "n/a", "Backend runtime image"),
    ("Hosting (backend)",  "Hugging Face Spaces (Docker SDK, free CPU)", "n/a", "Public HTTPS endpoint"),
    ("Hosting (frontend)", "Local Vite dev / Android APK / GitHub Pages (optional)", "n/a", "User-facing surface"),
    ("CI runner",          "GitHub Actions ubuntu-latest", "n/a", "APK build pipeline"),
]

API_SPEC = [
    # (method, path, request body, response, notes)
    ("GET",  "/api/health",  "", '{"status":"ok"}', "Liveness probe; used by HF Space healthcheck and the APK on startup"),
    ("POST", "/api/predict", '{"image":"data:image/jpeg;base64,..."}',
     '{"has_hand":bool, "label":str|null, "confidence":float}',
     "Top-1 prediction. has_hand=false means MediaPipe could not find a hand; label is null in that case."),
]

DATA_FLOW = [
    # (step, actor, action, payload)
    (1, "Browser / APK", "Capture frame via getUserMedia + canvas.toDataURL", "base64 JPEG, 640x480"),
    (2, "Browser / APK", "POST to /api/predict", "{image: ...}"),
    (3, "Flask API",     "Decode base64 -> Pillow -> NumPy BGR", "ndarray HxWx3"),
    (4, "Hand detector", "Run MediaPipe HandLandmarker -> 21 keypoints -> bbox", "(x, y, w, h)"),
    (5, "Flask API",     "Crop bbox + 20 px offset, resize 224x224, ImageNet-normalize", "tensor 1x3x224x224"),
    (6, "Classifier",    "ResNet-18 -> custom head -> softmax", "12-vector probabilities"),
    (7, "Flask API",     "argmax + lookup label name + round confidence", "JSON {has_hand, label, confidence}"),
    (8, "Browser / APK", "Render big emoji + word + confidence bar; speak word if >=80%", "DOM update + speechSynthesis.speak"),
    (9, "Browser / APK", "If confidence >=80%, append to recent-signs history (max 5, 3 s dedupe)", "UI strip update"),
]

RECOGNIZED_SIGNS = [
    ("milk",        "Milk",         "U+1F37C"),  # baby bottle
    ("eat",         "Eat",          "U+1F37D"),  # fork and knife with plate
    ("drink",       "Drink",        "U+1F964"),  # cup with straw
    ("down",        "Down",         "U+2B07"),   # down arrow
    ("mom",         "Mom",          "U+1F469"),  # woman
    ("mine",        "Mine",         "U+1F64B"),  # raised hand
    ("potty",       "Potty",        "U+1F6BD"),  # toilet
    ("sorry",       "Sorry",        "U+1F622"),  # crying face
    ("i_love_you",  "I love you",   "U+2764"),   # red heart
    ("mad_grumpy",  "Mad / Grumpy", "U+1F620"),  # angry face
    ("frustrated",  "Frustrated",   "U+1F624"),  # face with steam
    ("dont_know",   "Don't know",   "U+1F937"),  # shrug
]

MODEL_DETAILS = [
    ("Architecture",          "ResNet-18 (ImageNet-pretrained backbone) + custom classifier head"),
    ("Custom head",           "Linear(512 -> 256) -> ReLU -> Dropout(0.5) -> Linear(256 -> 12)"),
    ("Parameters",            "~11.31 M total"),
    ("Input shape",           "3 x 224 x 224, ImageNet normalize (mean=[0.485,0.456,0.406], std=[0.229,0.224,0.225])"),
    ("Output",                "12-class softmax (see Recognized Signs sheet)"),
    ("Source",                "RoxyDiya/BABY-SIGN-LANGUAGE-RECOGNITION (Sapienza University of Rome)"),
    ("Weights file",          "signlanguage_model.pth (44 MB), auto-downloaded on container boot"),
    ("Hand detector weights", "hand_landmarker.task (8 MB) from Google's MediaPipe model zoo"),
    ("Inference cost",        "~150 ms per frame on HF Spaces free CPU"),
]

DEPLOYMENT_TOPOLOGY = [
    # (environment, service, URL/path, notes)
    ("Production backend", "Hugging Face Space (Docker)",
     "https://codercarry-baby-sign-language-webapp.hf.space", "Public HTTPS, free tier"),
    ("Production backend (alt)", "Page",
     "https://huggingface.co/spaces/CoderCarry/baby-sign-language-webapp", "Build logs + git remote"),
    ("Source repo", "GitHub",
     "https://github.com/RahmanBhuiyan/baby-sign-language-webapp", "Code + GitHub Actions workflows"),
    ("APK artifact", "GitHub Actions",
     "Actions tab -> Build Android APK run -> Artifacts", "Debug-signed, ~4 MB, regenerated on push"),
    ("Local dev frontend", "Vite dev server",
     "http://localhost:5173", "npm run dev in frontend/"),
    ("Local dev backend",  "Flask dev server",
     "http://localhost:5000",  "py -3.13 backend/app.py"),
    ("CI variable", "VITE_API_BASE",
     "set in GitHub repo Settings -> Actions -> Variables", "Baked into APK at build time"),
]

FILE_MANIFEST = [
    # (path, purpose)
    ("backend/app.py",                       "Flask routes: /api/health, /api/predict, /(SPA fallback)"),
    ("backend/model_loader.py",              "Singleton ResNet18CustomClassifier + auto-download .pth"),
    ("backend/hand_detector.py",             "MediaPipe Tasks HandLandmarker singleton + auto-download .task"),
    ("backend/inference.py",                 "predict(b64_jpeg) -> {has_hand, label, confidence}"),
    ("backend/requirements.txt",             "Python deps incl. gunicorn"),
    ("backend/Dockerfile",                   "HF Spaces image: non-root user, libgles2, gunicorn"),
    ("backend/.dockerignore",                "Excludes __pycache__ and models/ from image"),
    ("backend/README.md",                    "HF Spaces frontmatter (sdk: docker, app_port: 7860)"),
    ("frontend/src/App.vue",                 "Top-level layout, wires composables to components"),
    ("frontend/src/components/CameraView.vue","getUserMedia + 200 ms capture loop"),
    ("frontend/src/components/LabelCard.vue","Big emoji + label + colored confidence bar"),
    ("frontend/src/components/HistoryStrip.vue","Last 5 signs with relative time"),
    ("frontend/src/components/SpeakToggle.vue","Voice on/off pill"),
    ("frontend/src/composables/usePredictor.js","POST to /api/predict with in-flight guard"),
    ("frontend/src/composables/useSpeech.js", "Web Speech API + 2 s repeat-suppress"),
    ("frontend/src/composables/useHistory.js","Ring buffer of 5, 3 s dedupe"),
    ("frontend/src/constants/labels.js",     "Label -> emoji + pretty name mapping; HIGH_CONF=0.80"),
    ("frontend/vite.config.js",              "Vite + /api proxy to :5000"),
    ("frontend/capacitor.config.json",       "appId: com.rahmanbhuiyan.babysignhelper, webDir: dist"),
    ("frontend/package.json",                "npm scripts + Capacitor deps"),
    (".github/workflows/android.yml",        "Builds APK on every push + workflow_dispatch"),
    ("docs/*.png",                           "README slide images (hero, architecture, features, signs grid)"),
    ("system-design/generate.py",            "This script - produces the PDF and Excel"),
]

CONFIDENCE_RULES = [
    (">= 80%", "Green bar; word is spoken aloud; chip is added to recent-signs history"),
    ("50-79%", "Yellow bar; label shown but not spoken; not added to history"),
    ("< 50%",  "Predictions hidden; UI shows 'Show me a sign'"),
    ("No hand detected", "UI shows 'Looking for hands...'; backend returns has_hand=false"),
    ("Repeat suppression (speech)", "Same word not spoken twice within 2 s"),
    ("Repeat suppression (history)","Same word not added to history twice within 3 s"),
]

FUTURE_WORK = [
    ("Serve frontend from HF too",
     "Multi-stage Dockerfile builds the Vue app and Flask serves it, so codercarry-baby-sign-language-webapp.hf.space/ shows the UI instead of the placeholder 503."),
    ("On-device TFLite",
     "Convert ResNet-18 to TFLite; run inference inside the APK so the app works offline and feels instant."),
    ("Play Store release",
     "Generate keystore, build signed AAB in CI, write privacy policy, fill Play Console listing."),
    ("Persistent paid backend",
     "Move HF free-tier backend to Render/Fly (~$5/mo) to remove cold-start delay."),
    ("More signs",
     "Retrain with additional baby-sign-language classes (e.g. 'more', 'all done', 'thank you')."),
    ("Authentication / rate limiting",
     "Add an API key check on /api/predict if the app gets shared publicly."),
]


# ----------------------------------------------------------------------
# PDF generation
# ----------------------------------------------------------------------
def build_pdf() -> None:
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"],
                        textColor=colors.HexColor("#2D2D44"), spaceAfter=8)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"],
                        textColor=colors.HexColor("#FF5FA3"), spaceBefore=14, spaceAfter=6)
    body = ParagraphStyle("body", parent=styles["BodyText"],
                          fontSize=10.5, leading=14, alignment=4)  # justify
    cover_title = ParagraphStyle("cover_title", parent=styles["Title"],
                                 fontSize=36, textColor=colors.HexColor("#2D2D44"),
                                 spaceAfter=14, alignment=1)
    cover_sub = ParagraphStyle("cover_sub", parent=styles["Heading2"],
                               fontSize=18, textColor=colors.HexColor("#6E6E8C"),
                               spaceAfter=20, alignment=1)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"],
                                fontSize=11, leading=16, alignment=1,
                                textColor=colors.HexColor("#555"))

    story = []

    # --- Cover ---
    story.append(Spacer(1, 4 * cm))
    hero = os.path.join(DOCS, "hero.png")
    if os.path.isfile(hero):
        story.append(Image(hero, width=16 * cm, height=6 * cm))
        story.append(Spacer(1, 1 * cm))
    story.append(Paragraph(META["title"], cover_title))
    story.append(Paragraph(META["subtitle"], cover_sub))
    story.append(Spacer(1, 2 * cm))
    story.append(Paragraph(f"Version {META['version']}", meta_style))
    story.append(Paragraph(META["date"], meta_style))
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(f"Owner: {META['owner']}", meta_style))
    story.append(Paragraph(META["repo"], meta_style))
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(f"Built on top of {META['based_on']}", meta_style))
    story.append(PageBreak())

    # --- 1. Overview ---
    story.append(Paragraph("1. Overview", h1))
    story.append(Paragraph(OVERVIEW, body))
    story.append(Spacer(1, 0.4 * cm))

    # --- 2. Architecture ---
    story.append(Paragraph("2. Architecture", h1))
    arch = os.path.join(DOCS, "architecture.png")
    if os.path.isfile(arch):
        story.append(Image(arch, width=17 * cm, height=7 * cm))
    story.append(Paragraph(
        "Every 200 ms the browser snaps a frame from the device camera, encodes "
        "it as a base64 JPEG, and POSTs it to <font color='#FF5FA3'>/api/predict</font>. "
        "The Flask backend decodes the image, runs MediaPipe HandLandmarker to "
        "find the hand bounding box, crops with a 20 px offset, resizes to "
        "224x224, normalizes for ImageNet, and runs the ResNet-18 classifier. "
        "The top-1 label and confidence return as JSON. The frontend shows a big "
        "emoji + word card, speaks the word through the Web Speech API when "
        "confidence is high, and keeps a rolling history of the last five signs.",
        body))
    story.append(PageBreak())

    # --- 3. Components ---
    story.append(Paragraph("3. Components", h1))
    story.append(_table(
        [["Component", "Role", "Technology", "Runtime location"]] +
        [list(r) for r in COMPONENTS],
        col_widths=[3.4 * cm, 6.0 * cm, 4.0 * cm, 4.0 * cm]))
    story.append(PageBreak())

    # --- 4. Tech stack ---
    story.append(Paragraph("4. Tech stack", h1))
    story.append(_table(
        [["Layer", "Technology", "Version", "Purpose"]] +
        [list(r) for r in TECH_STACK],
        col_widths=[4.0 * cm, 5.0 * cm, 2.4 * cm, 6.0 * cm]))
    story.append(PageBreak())

    # --- 5. API spec ---
    story.append(Paragraph("5. API specification", h1))
    story.append(_table(
        [["Method", "Path", "Request body", "Response", "Notes"]] +
        [list(r) for r in API_SPEC],
        col_widths=[1.6 * cm, 2.8 * cm, 4.6 * cm, 4.6 * cm, 3.8 * cm]))

    # --- 6. Data flow ---
    story.append(Paragraph("6. Data flow (per inference)", h2))
    story.append(_table(
        [["Step", "Actor", "Action", "Payload / shape"]] +
        [[str(s), a, b, c] for (s, a, b, c) in DATA_FLOW],
        col_widths=[1.2 * cm, 3.2 * cm, 8.6 * cm, 4.4 * cm]))
    story.append(PageBreak())

    # --- 7. ML model details ---
    story.append(Paragraph("7. ML model details", h1))
    story.append(_table(
        [["Property", "Value"]] +
        [list(r) for r in MODEL_DETAILS],
        col_widths=[4.5 * cm, 13.0 * cm]))

    # --- 8. Recognized signs ---
    story.append(Paragraph("8. Recognized signs", h2))
    signs = os.path.join(DOCS, "signs-grid.png")
    if os.path.isfile(signs):
        story.append(Image(signs, width=17 * cm, height=7.5 * cm))
    story.append(PageBreak())

    # --- 9. Confidence rules ---
    story.append(Paragraph("9. Confidence rules", h1))
    story.append(_table(
        [["Condition", "UI behaviour"]] +
        [list(r) for r in CONFIDENCE_RULES],
        col_widths=[5.0 * cm, 12.5 * cm]))

    # --- 10. Deployment topology ---
    story.append(Paragraph("10. Deployment topology", h2))
    story.append(_table(
        [["Environment", "Service", "URL / path", "Notes"]] +
        [list(r) for r in DEPLOYMENT_TOPOLOGY],
        col_widths=[3.4 * cm, 3.4 * cm, 6.5 * cm, 4.2 * cm]))
    story.append(PageBreak())

    # --- 11. File manifest ---
    story.append(Paragraph("11. File manifest (key paths)", h1))
    story.append(_table(
        [["Path", "Purpose"]] +
        [list(r) for r in FILE_MANIFEST],
        col_widths=[6.5 * cm, 11.0 * cm]))
    story.append(PageBreak())

    # --- 12. Future work ---
    story.append(Paragraph("12. Future work", h1))
    story.append(_table(
        [["Item", "Description"]] +
        [list(r) for r in FUTURE_WORK],
        col_widths=[4.5 * cm, 13.0 * cm]))

    doc = SimpleDocTemplate(
        PDF_OUT, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=f"{META['title']} - {META['subtitle']}",
        author=META["owner"],
    )
    doc.build(story)
    print(f"  PDF  -> {PDF_OUT}")


def _table(data, col_widths):
    """Build a styled platypus Table, wrapping cell text as Paragraphs so
    long strings don't overflow the page."""
    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=styles["BodyText"],
                          fontSize=8.5, leading=11)
    head = ParagraphStyle("head", parent=styles["BodyText"],
                          fontSize=9, leading=11, textColor=colors.white,
                          fontName="Helvetica-Bold")

    wrapped = []
    for r_idx, row in enumerate(data):
        wrapped.append([Paragraph(str(c), head if r_idx == 0 else cell) for c in row])

    t = Table(wrapped, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF7EB6")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
            [colors.HexColor("#FFFFFF"), colors.HexColor("#FCF6FF")]),
        ("BOX",        (0, 0), (-1, -1), 0.5, colors.HexColor("#EADBF0")),
        ("INNERGRID",  (0, 0), (-1, -1), 0.25, colors.HexColor("#EADBF0")),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
    ]))
    return t


# ----------------------------------------------------------------------
# Excel generation
# ----------------------------------------------------------------------
PINK   = "FF7EB6"
LAVNDR = "C7AFFF"
LIGHT  = "FCF6FF"
INK    = "2D2D44"
WHITE  = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=PINK)
ALT_FILL    = PatternFill("solid", fgColor=LIGHT)
HEAD_FONT   = Font(bold=True, color=WHITE, size=11)
BODY_FONT   = Font(color=INK, size=10)
TITLE_FONT  = Font(bold=True, color=INK, size=18)
SUB_FONT    = Font(italic=True, color="6E6E8C", size=11)

THIN = Side(border_style="thin", color="EADBF0")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autosize(ws, max_widths):
    for col_idx, max_w in enumerate(max_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max_w


def _write_table(ws, start_row, headers, rows, col_widths):
    """Write a header row + data rows starting at `start_row`. Returns the
    row index just after the table."""
    # header
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
    ws.row_dimensions[start_row].height = 24

    # body
    for r_offset, row in enumerate(rows):
        r = start_row + 1 + r_offset
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
            if r_offset % 2 == 1:
                cell.fill = ALT_FILL

    for col_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + 1 + len(rows)


def build_xlsx() -> None:
    wb = Workbook()

    # --- Overview sheet ---
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = META["title"]
    ws["A1"].font = TITLE_FONT
    ws["A2"] = META["subtitle"]
    ws["A2"].font = SUB_FONT
    ws.row_dimensions[1].height = 28

    info = [
        ("Version",  META["version"]),
        ("Date",     META["date"]),
        ("Owner",    META["owner"]),
        ("Repo",     META["repo"]),
        ("Based on", META["based_on"]),
        ("Backend",  "https://codercarry-baby-sign-language-webapp.hf.space"),
    ]
    for i, (k, v) in enumerate(info, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, color=INK)
        ws.cell(row=i, column=2, value=v).font = BODY_FONT

    ws.cell(row=4 + len(info) + 1, column=1, value="Description").font = Font(bold=True, color=INK)
    desc = ws.cell(row=4 + len(info) + 2, column=1, value=OVERVIEW)
    desc.font = BODY_FONT
    desc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=4 + len(info) + 2, start_column=1,
                   end_row=4 + len(info) + 8, end_column=5)
    _autosize(ws, [16, 60, 14, 14, 14])

    # --- Components ---
    ws = wb.create_sheet("Components")
    _write_table(ws, 1,
                 ["Component", "Role", "Technology", "Runtime location"],
                 COMPONENTS,
                 [22, 60, 32, 32])

    # --- Tech stack ---
    ws = wb.create_sheet("Tech stack")
    _write_table(ws, 1,
                 ["Layer", "Technology", "Version", "Purpose"],
                 TECH_STACK,
                 [28, 32, 16, 48])

    # --- API ---
    ws = wb.create_sheet("API")
    _write_table(ws, 1,
                 ["Method", "Path", "Request body", "Response", "Notes"],
                 API_SPEC,
                 [10, 18, 36, 36, 38])

    # --- Data flow ---
    ws = wb.create_sheet("Data flow")
    _write_table(ws, 1,
                 ["Step", "Actor", "Action", "Payload / shape"],
                 DATA_FLOW,
                 [6, 18, 60, 26])

    # --- Recognized signs ---
    ws = wb.create_sheet("Recognized signs")
    _write_table(ws, 1,
                 ["Label (model class)", "Pretty name", "Unicode codepoint"],
                 RECOGNIZED_SIGNS,
                 [22, 22, 22])

    # --- Model details ---
    ws = wb.create_sheet("Model details")
    _write_table(ws, 1,
                 ["Property", "Value"],
                 MODEL_DETAILS,
                 [26, 80])

    # --- Confidence rules ---
    ws = wb.create_sheet("Confidence rules")
    _write_table(ws, 1,
                 ["Condition", "UI behaviour"],
                 CONFIDENCE_RULES,
                 [30, 70])

    # --- Deployment ---
    ws = wb.create_sheet("Deployment")
    _write_table(ws, 1,
                 ["Environment", "Service", "URL / path", "Notes"],
                 DEPLOYMENT_TOPOLOGY,
                 [24, 26, 60, 36])

    # --- File manifest ---
    ws = wb.create_sheet("File manifest")
    _write_table(ws, 1,
                 ["Path", "Purpose"],
                 FILE_MANIFEST,
                 [44, 70])

    # --- Future work ---
    ws = wb.create_sheet("Future work")
    _write_table(ws, 1,
                 ["Item", "Description"],
                 FUTURE_WORK,
                 [28, 80])

    wb.save(XLSX_OUT)
    print(f"  XLSX -> {XLSX_OUT}")


# ----------------------------------------------------------------------
if __name__ == "__main__":
    print("Generating system design documents...")
    build_pdf()
    build_xlsx()
    print("Done.")
